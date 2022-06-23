# Author:      Jacob Moore
# Date:        5/26/22
# Description: This program...

from openpyxl import load_workbook
wb = load_workbook('ts_data.xlsx')

print(wb.sheetnames)