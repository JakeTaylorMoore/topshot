## DriveIndexer.py
## Created : April 6th, 2020
## Copyright Howard Leiner
## 
## Create an excel file to index a set of files/directories to aid ingestion into Avid

import os
#import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

SA_EXT = ['.PDF', '.DOCX', '.XLSX', '.HEIC', '.DOC', '.JPG', '.JPEG', '.TIF', '.PNG']
AA_EXT = ['.MP3', '.AIFF', '.WAV', '.WMA']
FA_EXT = ['.M4V', '.MP4', '.MOV', '.MKV', '.WEBM']

def writeHeader(sheet):
    sheet["A1"] = "#"
    sheet["B1"] = "Filename"
    sheet["C1"] = "Category"
    sheet["D1"] = "Avid File Name"
    sheet["E1"] = "Archive OX ID"
    sheet["F1"] = "Media Type"
    sheet["G1"] = "Source Code"
    sheet["H1"] = "Size(GB)"
    sheet["I1"] = "Size(Bytes)"
    sheet["J1"] = "Notes"
    sheet["K1"] = "Copyright/Original Source?"
    sheet["L1"] = "Description"
    sheet["M1"] = "Location"
    sheet["N1"] = "Archival Date"
    sheet["O1"] = "Duration"
    sheet["P1"] = "Image Aspect Ratio"
    sheet["Q1"] = "Detailed Description"
    sheet["R1"] = "Path"
    sheet["S1"] = "Dir1"
    sheet["T1"] = "Dir2"
    sheet["U1"] = "Original File Type"
    sheet["V1"] = "New File Type"
    sheet["W1"] = "Archival Ox #"
    sheet["X1"] = "NOTES FROM EDIT"

    for col in range(1, 25):
        cell = sheet.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fgColor="BFBFBF", fill_type = "solid")

def writeRow(sheet, i, filename, filetype, uniqueid, mediatype, sourcecode, avidfilename, dir1, dir2, path, size, sizegb, note):
    sheet["A"+str(i)] = i-1
    sheet["B"+str(i)] = filename
    sheet["U"+str(i)] = filetype
    sheet["W"+str(i)] = uniqueid
    sheet["F"+str(i)] = mediatype
    sheet["G"+str(i)] = sourcecode
    sheet["D"+str(i)] = avidfilename
    sheet["S"+str(i)] = dir1
    sheet["T"+str(i)] = dir2
    sheet["R"+str(i)] = path
    sheet["I"+str(i)] = size
    sheet["H"+str(i)] = sizegb
    sheet["J"+str(i)] = note

def getMediaType(ext):
    if (ext in SA_EXT):
        return "SA"
    elif (ext in AA_EXT):
        return "AA"
    elif (ext in FA_EXT):
        return "FA"
    else:
        return "??"

# create my excel file
workbook = Workbook()
sheet = workbook.active
writeHeader(sheet)

counter = 2
excelRow = 2
for root, dirs, files in os.walk("."):
    for name in files:
        if (not name.startswith('.')):
            pre_ext, ext = os.path.splitext(name)
            fileWithPath = os.path.join(root, name)
            sizeGB = os.stat(fileWithPath).st_size
            size = round(sizeGB/1000000000, 4)

            print(root)
            if (len(root) > 0):
                parts = root.split("/")

            if (len(parts) == 2):
                dir1 = parts[1]
                dir2 = ""
            elif (len(parts) >= 3):
                dir1 = parts[1]
                dir2 = parts[2]
            else:
                dir1 = ""
                dir2 = ""

            mediaType = getMediaType(ext.upper())

            r = str(excelRow)
            avidFileName = '="SRYVN_" & F' + r + ' & "_" & W' + r + ' & "_" & G' + r + ' & U' + r
            # SRYVN_<mediatype>_<id>_<sourcecode>.<ext>
            #print(avidFileName)

            writeRow (sheet, excelRow, name, ext, "XXXXX", mediaType, "TBD", avidFileName, dir1, dir2, fileWithPath, sizeGB, size, "")
            excelRow = excelRow + 1

        counter = counter + 1

print ('Total # of files', counter-1)
print ('Total # of files processed', excelRow-1)

workbook.save(filename="index.xlsx")

# reread my excel file and see if the files exist
#workbook2 = load_workbook(filename="index.xlsx")
#sheet = workbook2.active

#i = 1
#while (i < counter):
#    cell = sheet.cell(row = i, column = 2)
#    if (not os.path.exists(cell.value)):
#        print (cell.value)
#    i = i + 1


